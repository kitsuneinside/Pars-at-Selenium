from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
import undetected_chromedriver as uc
import openpyxl



def take_phone_name(page):
    try:
        return page.find_element(By.XPATH, '//h1').text
    except NoSuchElementException:
        return None


def take_phone_color(page):
    try:
        color = page.find_element(By.XPATH, "(//span[contains(@class, 'font-bold')])[1]").text.strip()
        if color:
            return color
        return None
    except NoSuchElementException:
        return None


def take_count_capacity(page):
    try:
        return page.find_element(By.XPATH, '//p[span[contains(text(), "Вбудована пам")]]/span[@class="font-bold"]').text
    except NoSuchElementException:
        return None


def take_phone_default_price(page):
    try:
        return page.find_element(By.XPATH, '//p[@class="product-price__small"]').text.strip()
    except NoSuchElementException:
        return None


def take_phone_price_with_discount(page):
    try:
        return page.find_element(By.XPATH, '//p[@class="product-price__big text-2xl font-bold leading-none '
                                           'product-price__big-color-red"]').text.strip()
    except NoSuchElementException:
        return None


def take_phone_price_without_discount(page):
    try:
        return page.find_element(By.XPATH, '//p[@class="product-price__big text-2xl font-bold leading-none"]').text
    except NoSuchElementException:
        return None


def take_phone_id(page):
    try:
        id = page.find_element(By.XPATH, "//span[contains(text(),'Код:')]")
        id_text = id.text
        return id_text.replace("Код:", "").strip()
    except NoSuchElementException:
        return None


def take_phone_count_feedback(page):
    try:
        button = page.find_element(
            By.XPATH,
            '//button[contains(@class, "comments-tabs__button") and contains(@class, "comments-tabs__active")]'
        )
        return button.get_attribute("innerText").strip()
    except NoSuchElementException:
        return None


def take_phone_diagonal(page):
    try:
        return page.find_element(By.XPATH, "//dt[span[normalize-space(text())='Діагональ екрана']]"
                                           "/following-sibling::dd[1]").text
    except NoSuchElementException:
        return None


def take_phone_resolution(page):
    try:
        return page.find_element(By.XPATH, "//dt[span[normalize-space(text())='Роздільна здатність дисплея']]"
                                           "/following-sibling::dd[1]").text
    except NoSuchElementException:
        return None


def take_phone_seria(page):
    try:
        return page.find_element(By.XPATH, "//dt[span[normalize-space(text())='Серія']]/following-sibling::dd[1]").text
    except NoSuchElementException:
        return None


def click_at_all_characteristics(page):
    try:
        return page.find_element(By.XPATH, "//rz-indexed-link[a[normalize-space(text())='Характеристики']]")
    except AttributeError:
        return None


def click_at_all_about_phone(page):
    try:
        return page.find_element(By.XPATH, "//rz-indexed-link[a[normalize-space(text())='Усе про товар']]")
    except AttributeError:
        return None


def take_all_characteristics(page):
    try:
        dts = page.find_elements(By.XPATH, "//rz-product-characteristics-list//dt")
        dds = page.find_elements(By.XPATH, "//rz-product-characteristics-list//dd")

        characteristics = {}
        for dt, dd in zip(dts, dds):
            name = dt.text.strip()
            value = dd.text.strip()
            characteristics[name] = value
        return characteristics
    except NoSuchElementException:
        return None


def take_all_photo(page):
    try:
        photo_links = page.find_elements(By.XPATH, '//rz-gallery-main-content-image//img')
        photos = []
        for links in photo_links:
            src = links.get_attribute('src')
            if src and src not in photos:
                photos.append(src)
        return photos
    except NoSuchElementException:
        return None


def take_seller(page):
    try:
        seller = page.find_element(By.XPATH, '//rz-seller-title//img[@class="seller-logo"]').get_attribute('alt')
        return seller
    except AttributeError:
        try:
            brand = page.find_element(By.XPATH, '//button[contains(@class, "product-producer")]//strong').text.strip()
            return brand
        except AttributeError:
            return None


def take_price(page):
    price = {}
    try:
        price['default'] = take_phone_price_without_discount(page)
        if price['default'] is None:
            price["small"] = take_phone_default_price(page)
            price['big'] = take_phone_price_with_discount(page)
        return price
    except NoSuchElementException:
        return None



def move_to_phone_page(page):
    wait = WebDriverWait(page, 15)
    action = ActionChains(page)
    try:
        page.get('https://rozetka.com.ua/')
        name_input = wait.until(EC.presence_of_element_located((By.XPATH, '//input')))
        action.move_to_element(name_input).perform()
        name_input.send_keys('Apple iPhone 15 128GB Black')
        find_phone = wait.until(EC.element_to_be_clickable((By.XPATH, '//button[contains(@class, '
                                                                      '"search-form__submit")]')))
        action.move_to_element(find_phone).perform()
        find_phone.click()
        wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'goods-tile__picture')))
        return page
    except AttributeError:
        return None


def move_to_phone(page):
    try:
        wait = WebDriverWait(page, 15)
        first_item = wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@class="goods-tile__picture"]')))
        first_item.click()
        wait.until(EC.presence_of_element_located((By.XPATH, '//h1')))
        return page
    except AttributeError:
        return None


def take_short_characteristics(page):
    try:
        photo_links = take_all_photo(page)
        price = take_price(page)
        ch = {
            "Назва": take_phone_name,
            "Колір": take_phone_color,
            "Памʼять": take_count_capacity,
            "Ціна повна": lambda *args: price['default'] if price.get('default') is not None else price.get('small'),
            "Ціна зі знижкою": lambda *args: price.get("big"),
            "ID": take_phone_id,
            "К-сть відгуків": take_phone_count_feedback,
            "Продавець": take_seller,
            "Серія": take_phone_seria,
            "Діагональ": take_phone_diagonal,
            "Роздільна здатність": take_phone_resolution,
            "Фото 1": lambda *_: photo_links[0] if len(photo_links) > 0 else None,
            "Фото 2": lambda *_: photo_links[1] if len(photo_links) > 1 else None,
            "Фото 3": lambda *_: photo_links[2] if len(photo_links) > 2 else None,
            "Фото 4": lambda *_: photo_links[3] if len(photo_links) > 3 else None,
            "Фото 5": lambda *_: photo_links[4] if len(photo_links) > 4 else None,
            "Фото 6": lambda *_: photo_links[5] if len(photo_links) > 5 else None,
            "Фото 7": lambda *_: photo_links[6] if len(photo_links) > 6 else None,
            "Фото 8": lambda *_: photo_links[7] if len(photo_links) > 7 else None,
            "Фото 9": lambda *_: photo_links[8] if len(photo_links) > 8 else None,
            "Фото 10": lambda *_: photo_links[9] if len(photo_links) > 9 else None,
        }
        r = {}
        for i, i2 in ch.items():
            r[i] = i2(page)
        return r
    except AttributeError:
        return None


def take_all_data(page):
    try:
        all_data = {}
        a = ActionChains(page)
        move = click_at_all_about_phone(page)
        a.move_to_element(move).perform()
        move.click()
        all_data['main'] = take_short_characteristics(page)
        move = click_at_all_characteristics(page)
        a.move_to_element(move).perform()
        move.click()
        all_data['second'] = take_all_characteristics(page)
        return all_data
    except AttributeError:
        return None


def print_all_data(data):
    for i, k in data.items():
        print(i)
        for key, value in k.items():
            print(f"{key}: {value}")


def write_in_exel(phone_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Phone Data'
    headers = [phone_data['main'].get("", i) for i in phone_data['main']]
    row = [phone_data['main'].get(h, "") for h in headers]
    ws.append(headers)
    ws.append(row)
    start_row = ws.max_row + 2
    ws.cell(row=start_row, column=1, value="Додаткові характеристики")
    for i, (k, v) in enumerate(phone_data['second'].items(), start=start_row + 1):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
    wb.save("iphone_data.xlsx")


def main():
    options = uc.ChromeOptions()
    options.binary_location = "/Users/oleh/Desktop/Google Chrome.app/Contents/MacOS/Google Chrome"
    driver = uc.Chrome(options=options, use_subprocess=False)
    action = ActionChains(driver)
    move_to_phone_page(driver)
    move_to_phone(driver)
    move_to = click_at_all_characteristics(driver)
    action.move_to_element(move_to).perform()
    move_to.click()
    all_data = take_all_data(driver)
    print_all_data(all_data)
    write_in_exel(all_data)
    driver.quit()


if __name__ == "__main__":
    main()
